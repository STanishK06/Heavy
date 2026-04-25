"""
routes/inquiries.py â€” Inquiries (HeavyLift CRM)
Features: auto-followup notification, optional refs, offer linkage, WhatsApp, Excel export
"""
from datetime import date, timedelta
import io
import urllib.parse

from flask import Blueprint, current_app, flash, jsonify, make_response, redirect, render_template, request, session, url_for

from database import close_db, get_db
from routes.auth import login_required, role_required
from routes.notifications import create_notification
from validation import clean_choice, clean_optional_text, clean_text, parse_optional_int

inquiries_bp = Blueprint("inquiries", __name__, url_prefix="/inquiries")


def _scope(role, loc_id, uid):
    base = (
        "FROM inquiries i "
        "LEFT JOIN locations l ON i.location_id=l.id "
        "LEFT JOIN courses c ON i.course_id=c.id "
        "LEFT JOIN offers o ON i.offer_id=o.id "
        "WHERE 1=1"
    )
    params = []
    if role == "teacher" and loc_id:
        base += " AND i.location_id=%s"
        params.append(loc_id)
    return base, params


def _fetch_inquiry(cur, iid, role, loc_id, with_joins=False):
    select_sql = "i.*"
    joins_sql = ""
    if with_joins:
        select_sql = "i.*, l.name AS location_name, c.name AS course_name"
        joins_sql = (
            " LEFT JOIN locations l ON i.location_id=l.id"
            " LEFT JOIN courses c ON i.course_id=c.id"
        )

    query = f"SELECT {select_sql} FROM inquiries i{joins_sql} WHERE i.id=%s"
    params = [iid]
    if role == "teacher" and loc_id:
        query += " AND i.location_id=%s"
        params.append(loc_id)
    cur.execute(query + ";", params)
    return cur.fetchone()


def _normalize_mobile(value):
    digits = "".join(ch for ch in (value or "") if ch.isdigit())
    if len(digits) == 12 and digits.startswith("91"):
        digits = digits[2:]
    if len(digits) != 10:
        raise ValueError("Mobile must be a valid 10-digit number.")
    return digits


def _normalize_optional_mobile(value, field_name):
    return clean_optional_text(value, field_name, max_length=20)


def _render_inquiry_form(*, inquiry, locations, courses, offers, defaults, action, form_data=None, form_error_popup=None):
    return render_template(
        "inquiries/form.html",
        inquiry=inquiry,
        locations=locations,
        courses=courses,
        offers=offers,
        defaults=defaults,
        action=action,
        form_data=form_data or {},
        form_error_popup=form_error_popup,
    )


def _parse_amount(value, field_name):
    raw = str(value or "0").replace(",", "").strip()
    try:
        amount = float(raw or 0)
    except ValueError as exc:
        raise ValueError(f"{field_name} must be a valid number.") from exc
    if amount < 0:
        raise ValueError(f"{field_name} cannot be negative.")
    return amount


def _parse_date(value, field_name, required=False):
    raw = (value or "").strip()
    if not raw:
        if required:
            raise ValueError(f"{field_name} is required.")
        return None
    try:
        return date.fromisoformat(raw)
    except ValueError as exc:
        raise ValueError(f"{field_name} must be a valid date.") from exc


def _validate_inquiry_form(form, fees_total):
    name = clean_text(form.get("name"), "Name", required=True, max_length=120)
    mobile = _normalize_mobile(form.get("mobile"))
    inquiry_date = _parse_date(form.get("inquiry_date"), "Inquiry date", required=True)
    followup_date = _parse_date(form.get("followup_date"), "Follow-up date")
    admission_date = _parse_date(form.get("admission_date"), "Admission date")
    fees_paid = _parse_amount(form.get("fees_paid", "0"), "Fees paid")
    gender = clean_choice(form.get("gender"), "Gender", {"Male", "Female", "Other"}, required=False)
    status = clean_choice(form.get("status", "Open"), "Status", {"Open", "Converted", "Closed"})
    city = clean_optional_text(form.get("city"), "City", max_length=80)
    state = clean_optional_text(form.get("state"), "State", max_length=80)
    ref1_name = clean_optional_text(form.get("ref1_name"), "Reference 1 name", max_length=100)
    ref2_name = clean_optional_text(form.get("ref2_name"), "Reference 2 name", max_length=100)
    ref3_name = clean_optional_text(form.get("ref3_name"), "Reference 3 name", max_length=100)
    ref1_mobile = _normalize_optional_mobile(form.get("ref1_mobile"), "Reference 1 mobile")
    ref2_mobile = _normalize_optional_mobile(form.get("ref2_mobile"), "Reference 2 mobile")
    ref3_mobile = _normalize_optional_mobile(form.get("ref3_mobile"), "Reference 3 mobile")

    if followup_date and followup_date < inquiry_date:
        raise ValueError("Follow-up date cannot be earlier than inquiry date.")
    if admission_date and admission_date < inquiry_date:
        raise ValueError("Admission date cannot be earlier than inquiry date.")
    if fees_paid > fees_total:
        raise ValueError("Fees paid cannot be greater than total fees.")

    return {
        "name": name,
        "mobile": mobile,
        "gender": gender,
        "status": status,
        "city": city,
        "state": state,
        "inquiry_date": inquiry_date.isoformat(),
        "followup_date": followup_date.isoformat() if followup_date else None,
        "admission_date": admission_date.isoformat() if admission_date else None,
        "fees_paid": fees_paid,
        "ref1_name": ref1_name,
        "ref1_mobile": ref1_mobile,
        "ref2_name": ref2_name,
        "ref2_mobile": ref2_mobile,
        "ref3_name": ref3_name,
        "ref3_mobile": ref3_mobile,
    }


def _calculate_total_fees(course_id, offer_id):
    if not course_id:
        return 0.0

    conn = get_db()
    cur = conn.cursor()
    try:
        cur.execute("SELECT fees FROM courses WHERE id=%s;", (course_id,))
        row = cur.fetchone()
        fees_total = float(row["fees"]) if row else 0.0
        if offer_id:
            cur.execute("SELECT discount_type,discount_value FROM offers WHERE id=%s;", (offer_id,))
            offer = cur.fetchone()
            if offer:
                if offer["discount_type"] == "percent":
                    fees_total -= fees_total * float(offer["discount_value"]) / 100
                else:
                    fees_total -= float(offer["discount_value"])
        return max(0.0, fees_total)
    finally:
        close_db(conn, commit=False)


@inquiries_bp.route("/")
@login_required
def index():
    role = session.get("role")
    loc_id = session.get("location_id")
    uid = session.get("user_id")
    raw_status = clean_text(request.args.get("status", ""), "Status", max_length=20)
    status_filter = raw_status if raw_status in {"", "Open", "Converted", "Closed"} else ""
    filters = {
        "name": clean_text(request.args.get("name", ""), "Name", max_length=120),
        "mobile": clean_text(request.args.get("mobile", ""), "Mobile", max_length=20),
        "location": clean_text(request.args.get("location", ""), "Location", max_length=100),
        "course": clean_text(request.args.get("course", ""), "Course", max_length=100),
        "status": status_filter,
        "date_from": clean_text(request.args.get("date_from", ""), "Date from", max_length=20),
        "date_to": clean_text(request.args.get("date_to", ""), "Date to", max_length=20),
        "last_days": clean_text(request.args.get("last_days", ""), "Last days", max_length=4),
    }
    sort_col = request.args.get("sort", "i.inquiry_date")
    sort_dir = request.args.get("dir", "desc")
    allowed = {"i.inquiry_date", "i.name", "i.mobile", "l.name", "c.name", "i.status", "i.followup_date"}
    if sort_col not in allowed:
        sort_col = "i.inquiry_date"
    if sort_dir not in ("asc", "desc"):
        sort_dir = "desc"

    base, params = _scope(role, loc_id, uid)
    if filters["name"]:
        base += " AND i.name ILIKE %s"
        params.append(f"%{filters['name']}%")
    if filters["mobile"]:
        base += " AND i.mobile ILIKE %s"
        params.append(f"%{filters['mobile']}%")
    if filters["location"]:
        base += " AND l.name ILIKE %s"
        params.append(f"%{filters['location']}%")
    if filters["course"]:
        base += " AND c.name ILIKE %s"
        params.append(f"%{filters['course']}%")
    if filters["status"]:
        base += " AND i.status=%s"
        params.append(filters["status"])
    if filters["last_days"]:
        try:
            base += " AND i.inquiry_date >= %s"
            params.append(date.today() - timedelta(days=int(filters["last_days"])))
        except ValueError:
            flash("Last X Days must be a valid number.", "warning")
    if filters["date_from"]:
        base += " AND i.inquiry_date >= %s"
        params.append(filters["date_from"])
    if filters["date_to"]:
        base += " AND i.inquiry_date <= %s"
        params.append(filters["date_to"])

    conn = get_db()
    cur = conn.cursor()
    cur.execute(f"SELECT i.*,l.name AS location_name,c.name AS course_name,o.name AS offer_name {base} ORDER BY {sort_col} {sort_dir};", params)
    inquiries = cur.fetchall()
    if role == "teacher" and loc_id:
        cur.execute("SELECT id,name FROM locations WHERE id=%s ORDER BY position,name;", (loc_id,))
        locations = cur.fetchall()
        cur.execute("SELECT id,name FROM courses WHERE location_id=%s ORDER BY position,name;", (loc_id,))
        courses = cur.fetchall()
    else:
        cur.execute("SELECT id,name FROM locations ORDER BY position,name;")
        locations = cur.fetchall()
        cur.execute("SELECT id,name FROM courses ORDER BY position,name;")
        courses = cur.fetchall()
    close_db(conn, commit=False)
    return render_template(
        "inquiries/index.html",
        inquiries=inquiries,
        locations=locations,
        courses=courses,
        filters=filters,
        sort_col=sort_col,
        sort_dir=sort_dir,
        today=date.today(),
    )


@inquiries_bp.route("/add", methods=["GET", "POST"])
@login_required
@role_required("admin", "developer", "teacher")
def add():
    role = session.get("role")
    assigned_loc_id = session.get("location_id")
    conn = get_db()
    cur = conn.cursor()
    if role == "teacher" and assigned_loc_id:
        cur.execute("SELECT id,name FROM locations WHERE id=%s ORDER BY position,name;", (assigned_loc_id,))
        locs = cur.fetchall()
        cur.execute("SELECT id,name,fees FROM courses WHERE location_id=%s ORDER BY position,name;", (assigned_loc_id,))
        courses_list = cur.fetchall()
    else:
        cur.execute("SELECT id,name FROM locations ORDER BY position,name;")
        locs = cur.fetchall()
        cur.execute("SELECT id,name,fees FROM courses ORDER BY position,name;")
        courses_list = cur.fetchall()
    if role == "teacher" and assigned_loc_id:
        cur.execute(
            "SELECT id,name,discount_type,discount_value FROM offers "
            "WHERE is_active=TRUE AND (valid_to IS NULL OR valid_to >= CURRENT_DATE) "
            "AND (location_id IS NULL OR location_id=%s) ORDER BY name;",
            (assigned_loc_id,),
        )
    else:
        cur.execute(
            "SELECT id,name,discount_type,discount_value FROM offers "
            "WHERE is_active=TRUE AND (valid_to IS NULL OR valid_to >= CURRENT_DATE) ORDER BY name;"
        )
    offers = cur.fetchall()
    close_db(conn, commit=False)

    defaults = {
        "inquiry_date": date.today().isoformat(),
        "followup_date": (date.today() + timedelta(days=10)).isoformat(),
    }

    if request.method == "POST":
        form = request.form
        try:
            location_id = parse_optional_int(form.get("location_id"), "Location")
            if role == "teacher" and assigned_loc_id:
                location_id = assigned_loc_id
            course_id = parse_optional_int(form.get("course_id"), "Course")
            offer_id = parse_optional_int(form.get("offer_id"), "Offer")
            fees_total = _calculate_total_fees(course_id, offer_id)
            cleaned = _validate_inquiry_form(form, fees_total)
            if role == "teacher" and assigned_loc_id:
                conn_check = get_db()
                cur_check = conn_check.cursor()
                try:
                    if course_id:
                        cur_check.execute("SELECT 1 FROM courses WHERE id=%s AND location_id=%s;", (course_id, assigned_loc_id))
                        if not cur_check.fetchone():
                            raise ValueError("You can only use courses from your assigned location.")
                    if offer_id:
                        cur_check.execute(
                            "SELECT 1 FROM offers WHERE id=%s AND (location_id IS NULL OR location_id=%s);",
                            (offer_id, assigned_loc_id),
                        )
                        if not cur_check.fetchone():
                            raise ValueError("You can only use offers available to your assigned location.")
                finally:
                    close_db(conn_check, commit=False)
            conn = get_db()
            cur = conn.cursor()
            cur.execute(
                """
                INSERT INTO inquiries
                  (name,gender,mobile,location_id,city,state,course_id,offer_id,
                   inquiry_date,followup_date,admission_date,status,fees_total,fees_paid,
                   ref1_name,ref1_mobile,ref2_name,ref2_mobile,ref3_name,ref3_mobile)
                VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s);
                """,
                (
                    cleaned["name"],
                    cleaned["gender"],
                    cleaned["mobile"],
                    location_id,
                    cleaned["city"],
                    cleaned["state"],
                    course_id,
                    offer_id,
                    cleaned["inquiry_date"],
                    cleaned["followup_date"],
                    cleaned["admission_date"],
                    cleaned["status"],
                    fees_total,
                    cleaned["fees_paid"],
                    cleaned["ref1_name"],
                    cleaned["ref1_mobile"],
                    cleaned["ref2_name"],
                    cleaned["ref2_mobile"],
                    cleaned["ref3_name"],
                    cleaned["ref3_mobile"],
                ),
            )
            close_db(conn)
            create_notification(
                f"Follow-up due: {cleaned['name']}",
                f"Follow-up scheduled on {cleaned['followup_date'] or 'N/A'} for {cleaned['name']} ({cleaned['mobile']}).",
                target_role="admin",
            )
            flash("Inquiry added.", "success")
            return redirect(url_for("inquiries.index"))
        except ValueError as exc:
            flash(str(exc), "danger")
            return _render_inquiry_form(
                inquiry=None,
                locations=locs,
                courses=courses_list,
                offers=offers,
                defaults=defaults,
                action="Add",
                form_data=request.form.to_dict(flat=True),
                form_error_popup=str(exc),
            )
        except Exception:
            current_app.logger.exception("Failed to add inquiry")
            message = "Unable to save inquiry right now."
            flash(message, "danger")
            return _render_inquiry_form(
                inquiry=None,
                locations=locs,
                courses=courses_list,
                offers=offers,
                defaults=defaults,
                action="Add",
                form_data=request.form.to_dict(flat=True),
                form_error_popup=message,
            )

    return _render_inquiry_form(
        inquiry=None,
        locations=locs,
        courses=courses_list,
        offers=offers,
        defaults=defaults,
        action="Add",
    )


@inquiries_bp.route("/<int:iid>/edit", methods=["GET", "POST"])
@login_required
@role_required("admin", "developer", "teacher")
def edit(iid):
    role = session.get("role")
    loc_id = session.get("location_id")
    conn = get_db()
    cur = conn.cursor()
    inquiry = _fetch_inquiry(cur, iid, role, loc_id)
    if role == "teacher" and loc_id:
        cur.execute("SELECT id,name FROM locations WHERE id=%s ORDER BY position,name;", (loc_id,))
        locs = cur.fetchall()
        cur.execute("SELECT id,name,fees FROM courses WHERE location_id=%s ORDER BY position,name;", (loc_id,))
        courses_list = cur.fetchall()
        cur.execute(
            "SELECT id,name,discount_type,discount_value FROM offers "
            "WHERE is_active=TRUE AND (location_id IS NULL OR location_id=%s) ORDER BY name;",
            (loc_id,),
        )
        offers = cur.fetchall()
    else:
        cur.execute("SELECT id,name FROM locations ORDER BY position,name;")
        locs = cur.fetchall()
        cur.execute("SELECT id,name,fees FROM courses ORDER BY position,name;")
        courses_list = cur.fetchall()
        cur.execute("SELECT id,name,discount_type,discount_value FROM offers WHERE is_active=TRUE ORDER BY name;")
        offers = cur.fetchall()
    close_db(conn, commit=False)
    if not inquiry:
        flash("Not found.", "danger")
        return redirect(url_for("inquiries.index"))

    if request.method == "POST":
        form = request.form
        try:
            fees_total = _parse_amount(form.get("fees_total", "0"), "Fees total")
            cleaned = _validate_inquiry_form(form, fees_total)
            location_id = parse_optional_int(form.get("location_id"), "Location")
            course_id = parse_optional_int(form.get("course_id"), "Course")
            offer_id = parse_optional_int(form.get("offer_id"), "Offer")
            conn = get_db()
            cur = conn.cursor()
            if not _fetch_inquiry(cur, iid, role, loc_id):
                close_db(conn, commit=False)
                flash("Not found.", "danger")
                return redirect(url_for("inquiries.index"))
            if role == "teacher" and loc_id:
                if course_id:
                    cur.execute("SELECT 1 FROM courses WHERE id=%s AND location_id=%s;", (course_id, loc_id))
                    if not cur.fetchone():
                        close_db(conn, commit=False)
                        raise ValueError("You can only use courses from your assigned location.")
                if offer_id:
                    cur.execute(
                        "SELECT 1 FROM offers WHERE id=%s AND (location_id IS NULL OR location_id=%s);",
                        (offer_id, loc_id),
                    )
                    if not cur.fetchone():
                        close_db(conn, commit=False)
                        raise ValueError("You can only use offers available to your assigned location.")
            cur.execute(
                """
                UPDATE inquiries SET
                  name=%s,gender=%s,mobile=%s,location_id=%s,city=%s,state=%s,
                  course_id=%s,offer_id=%s,inquiry_date=%s,followup_date=%s,
                  admission_date=%s,status=%s,fees_total=%s,fees_paid=%s,
                  ref1_name=%s,ref1_mobile=%s,ref2_name=%s,ref2_mobile=%s,
                  ref3_name=%s,ref3_mobile=%s
                WHERE id=%s;
                """,
                (
                    cleaned["name"],
                    cleaned["gender"],
                    cleaned["mobile"],
                    loc_id if role == "teacher" and loc_id else location_id,
                    cleaned["city"],
                    cleaned["state"],
                    course_id,
                    offer_id,
                    cleaned["inquiry_date"],
                    cleaned["followup_date"],
                    cleaned["admission_date"],
                    cleaned["status"],
                    fees_total,
                    cleaned["fees_paid"],
                    cleaned["ref1_name"],
                    cleaned["ref1_mobile"],
                    cleaned["ref2_name"],
                    cleaned["ref2_mobile"],
                    cleaned["ref3_name"],
                    cleaned["ref3_mobile"],
                    iid,
                ),
            )
            close_db(conn)
            flash("Updated.", "success")
            return redirect(url_for("inquiries.index"))
        except ValueError as exc:
            flash(str(exc), "danger")
            return _render_inquiry_form(
                inquiry=inquiry,
                locations=locs,
                courses=courses_list,
                offers=offers,
                defaults={},
                action="Edit",
                form_data=request.form.to_dict(flat=True),
                form_error_popup=str(exc),
            )
        except Exception:
            current_app.logger.exception("Failed to update inquiry %s", iid)
            message = "Unable to update inquiry right now."
            flash(message, "danger")
            return _render_inquiry_form(
                inquiry=inquiry,
                locations=locs,
                courses=courses_list,
                offers=offers,
                defaults={},
                action="Edit",
                form_data=request.form.to_dict(flat=True),
                form_error_popup=message,
            )

    return _render_inquiry_form(
        inquiry=inquiry,
        locations=locs,
        courses=courses_list,
        offers=offers,
        defaults={},
        action="Edit",
    )


@inquiries_bp.route("/<int:iid>/delete", methods=["POST"])
@login_required
@role_required("admin", "developer")
def delete(iid):
    conn = get_db()
    cur = conn.cursor()
    cur.execute("DELETE FROM inquiries WHERE id=%s;", (iid,))
    close_db(conn)
    flash("Deleted.", "success")
    return redirect(url_for("inquiries.index"))


@inquiries_bp.route("/<int:iid>/convert", methods=["POST"])
@login_required
@role_required("admin", "developer", "teacher")
def convert(iid):
    role = session.get("role")
    loc_id = session.get("location_id")
    conn = get_db()
    cur = conn.cursor()
    inquiry = _fetch_inquiry(cur, iid, role, loc_id)
    if not inquiry:
        close_db(conn, commit=False)
        flash("Not found.", "danger")
        return redirect(url_for("inquiries.index"))
    cur.execute(
        "UPDATE inquiries SET status='Converted', admission_date=COALESCE(admission_date,CURRENT_DATE) WHERE id=%s;",
        (iid,),
    )
    close_db(conn)
    create_notification(f"Admission: {inquiry['name']}", f"{inquiry['name']} has been converted to a student.", "admin")
    flash(f"{inquiry['name']} converted to student.", "success")
    return redirect(url_for("inquiries.index"))


@inquiries_bp.route("/<int:iid>/followup", methods=["GET", "POST"])
@login_required
@role_required("admin", "developer", "teacher")
def followup(iid):
    role = session.get("role")
    loc_id = session.get("location_id")
    conn = get_db()
    cur = conn.cursor()
    inquiry = _fetch_inquiry(cur, iid, role, loc_id, with_joins=True)
    if not inquiry:
        close_db(conn, commit=False)
        flash("Not found.", "danger")
        return redirect(url_for("inquiries.index"))

    if request.method == "POST":
        try:
            conversation = clean_optional_text(request.form.get("conversation", ""), "Conversation", max_length=4000, multiline=True)
            followup_date = _parse_date(request.form.get("followup_date"), "Follow-up date")
            status = clean_choice(request.form.get("status", inquiry["status"]), "Status", {"Open", "Converted", "Closed"})
            admission_date = _parse_date(request.form.get("admission_date"), "Admission date")
            if followup_date and followup_date < inquiry["inquiry_date"]:
                raise ValueError("Follow-up date cannot be earlier than inquiry date.")
            if admission_date and admission_date < inquiry["inquiry_date"]:
                raise ValueError("Admission date cannot be earlier than inquiry date.")
            cur.execute(
                "INSERT INTO followups (inquiry_id,conversation,followup_date,status) VALUES (%s,%s,%s,%s);",
                (iid, conversation, followup_date.isoformat() if followup_date else None, status),
            )
            cur.execute(
                "UPDATE inquiries SET status=%s,followup_date=%s,admission_date=COALESCE(%s::date,admission_date) WHERE id=%s;",
                (status, followup_date.isoformat() if followup_date else None, admission_date.isoformat() if admission_date else None, iid),
            )
            close_db(conn)
            if followup_date:
                create_notification(f"Next follow-up: {inquiry['name']}", f"Scheduled for {followup_date.isoformat()}.", "admin")
            flash("Follow-up saved.", "success")
            return redirect(url_for("inquiries.followup", iid=iid))
        except ValueError as exc:
            flash(str(exc), "danger")
        except Exception:
            current_app.logger.exception("Failed to save follow-up for inquiry %s", iid)
            flash("Unable to save follow-up right now.", "danger")

    cur.execute("SELECT * FROM followups WHERE inquiry_id=%s ORDER BY created_at DESC;", (iid,))
    followups = cur.fetchall()
    close_db(conn, commit=False)
    default_next = (date.today() + timedelta(days=7)).isoformat()
    return render_template("inquiries/followup.html", inquiry=inquiry, followups=followups, default_next=default_next)


@inquiries_bp.route("/<int:iid>/whatsapp-send", methods=["POST"])
@login_required
def send_whatsapp(iid):
    """Return wa.me link for direct WhatsApp send (or call API if configured)."""
    if not request.is_json:
        return jsonify({"ok": False, "msg": "JSON body required"}), 400
    role = session.get("role")
    loc_id = session.get("location_id")
    payload = request.get_json(silent=True) or {}
    conn = get_db()
    cur = conn.cursor()
    inquiry = _fetch_inquiry(cur, iid, role, loc_id)
    if not inquiry:
        close_db(conn, commit=False)
        return jsonify({"ok": False, "msg": "Not found"}), 404

    try:
        msg_id = parse_optional_int(payload.get("msg_id"), "Message template")
        msg_text = clean_optional_text(payload.get("message"), "Message", max_length=2000, multiline=True) or ""
    except ValueError as exc:
        close_db(conn, commit=False)
        return jsonify({"ok": False, "msg": str(exc)}), 400
    if msg_id:
        cur.execute("SELECT description FROM whatsapp_msgs WHERE id=%s;", (msg_id,))
        template = cur.fetchone()
        if template:
            msg_text = (template["description"] or "").replace("[NAME]", inquiry["name"]).replace("[MOBILE]", inquiry["mobile"])
    close_db(conn, commit=False)

    mobile = inquiry["mobile"].replace(" ", "").replace("-", "").replace("+", "")
    if not mobile.startswith("91"):
        mobile = "91" + mobile
    wa_url = f"https://wa.me/{mobile}?text={urllib.parse.quote(msg_text)}"
    return jsonify({"ok": True, "url": wa_url})


@inquiries_bp.route("/export")
@login_required
@role_required("admin", "developer", "teacher")
def export():
    import openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill

    role = session.get("role")
    loc_id = session.get("location_id")
    uid = session.get("user_id")
    base, params = _scope(role, loc_id, uid)
    conn = get_db()
    cur = conn.cursor()
    cur.execute(f"SELECT i.*,l.name AS location_name,c.name AS course_name {base} ORDER BY i.inquiry_date DESC;", params)
    rows = cur.fetchall()
    close_db(conn, commit=False)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Inquiries"
    headers = [
        "ID", "Name", "Gender", "Mobile", "Location", "City", "State", "Course",
        "Inquiry Date", "Followup Date", "Admission Date", "Status",
        "Fees Total", "Fees Paid", "Pending", "Ref1 Name", "Ref1 Mobile",
    ]
    hfill = PatternFill("solid", fgColor="F59E0B")
    hfont = Font(color="000000", bold=True)
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = hfill
        cell.font = hfont
        cell.alignment = Alignment(horizontal="center")
    for row_index, row in enumerate(rows, 2):
        pending = float(row.get("fees_total") or 0) - float(row.get("fees_paid") or 0)
        values = [
            row["id"], row["name"], row["gender"], row["mobile"], row["location_name"],
            row["city"], row["state"], row["course_name"],
            str(row["inquiry_date"]) if row["inquiry_date"] else "",
            str(row["followup_date"]) if row["followup_date"] else "",
            str(row["admission_date"]) if row["admission_date"] else "",
            row["status"], row.get("fees_total", 0), row.get("fees_paid", 0),
            pending, row.get("ref1_name", ""), row.get("ref1_mobile", ""),
        ]
        for col_index, value in enumerate(values, 1):
            ws.cell(row=row_index, column=col_index, value=value)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    response = make_response(buf.read())
    response.headers["Content-Disposition"] = "attachment; filename=inquiries.xlsx"
    response.headers["Content-Type"] = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    return response
